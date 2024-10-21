from django.shortcuts import render, redirect
from django.contrib.auth.views import (
    LoginView,
    PasswordResetView,
    PasswordChangeView,
    PasswordResetConfirmView,
)
from .forms import (
    RegistrationForm,
    UserLoginForm,
    UserPasswordResetForm,
    UserSetPasswordForm,
    UserPasswordChangeForm,
    StudentRegistrationForm,
    CoordinatorRegistrationForm,
    FacultyRegistrationForm,
)
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from .models import Project
from .forms import ProjectForm
from django.contrib.auth import get_user_model
from django.views.generic import ListView, CreateView, DetailView, UpdateView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.edit import DeleteView
from django.urls import reverse
from .models import ExaminerMidProjectPresentation, ExaminerMidProjectPresentationQA
from .forms import (
    ExaminerMidProjectPresentationForm,
    ExaminerMidProjectPresentationQAForm,
    ExaminerMidProjectPresentationQAFormSet,
)
from .forms import LogbookEntryFormSet

User = get_user_model()

# Create your views here.


@login_required
def update_profile(request):
    if request.method == "POST":
        user_id = request.user.id
        user = User.objects.get(pk=user_id)
        user.first_name = request.POST.get("first_name")
        user.last_name = request.POST.get("last_name")
        user.email = request.POST.get("email")
        user.about = request.POST.get("about")
        user.save()
        messages.success(request, "Your profile was successfully updated!")
        return redirect("update_profile")
    else:
        return render(request, "pages/profile.html")


# Pages
def index(request):
    return render(request, "pages/index.html")


def about_us(request):
    return render(request, "pages/about-us.html")


def contact_us(request):
    return render(request, "pages/contact-us.html")


def author(request):
    return render(request, "pages/author.html")


# Authentication


def register(request):
    return render(request, "accounts/sign-up-user.html")


def student_register(request):
    if request.method == "POST":
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            print("Account created successfully!")
            return redirect("/accounts/login")
        else:
            print("Registration failed!")
    else:
        form = StudentRegistrationForm()

    context = {"form": form}
    return render(request, "accounts/sign-up.html", context)


def coordinator_register(request):
    if request.method == "POST":
        form = CoordinatorRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            print("Account created successfully!")
            return redirect("/accounts/login")
        else:
            print("Registration failed!")
    else:
        form = CoordinatorRegistrationForm()

    context = {"form": form}
    return render(request, "accounts/sign-up.html", context)


def faculty_register(request):
    if request.method == "POST":
        form = FacultyRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            print("Account created successfully!")
            return redirect("/accounts/login")
        else:
            print("Registration failed!")
    else:
        form = FacultyRegistrationForm()

    context = {"form": form}
    return render(request, "accounts/sign-up.html", context)


class UserLoginView(LoginView):
    template_name = "accounts/sign-in.html"
    form_class = UserLoginForm


def logout_view(request):
    logout(request)
    return redirect("/accounts/login")


class UserPasswordResetView(PasswordResetView):
    template_name = "accounts/password_reset.html"
    form_class = UserPasswordResetForm


class UserPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = "accounts/password_reset_confirm.html"
    form_class = UserSetPasswordForm


class UserPasswordChangeView(PasswordChangeView):
    template_name = "accounts/password_change.html"
    form_class = UserPasswordChangeForm


class ProjectListView(LoginRequiredMixin, ListView):
    model = Project
    template_name = "coordinator/project_list.html"
    context_object_name = "projects"
    login_url = "/accounts/login/"
    redirect_field_name = "redirect_to"

    def get_queryset(self):
        # Customize the queryset to only show projects supervised by the current user
        return Project.objects.filter(coordinator=self.request.user)


class ProjectCreateView(CreateView):
    model = Project
    form_class = ProjectForm
    template_name = "coordinator/project_form.html"
    success_url = reverse_lazy("coordinator-project-list")

    def form_valid(self, form):
        # Coordinator creates the project without assigning a supervisor
        project = form.save(commit=False)
        project.coordinator = (
            self.request.user
        )  # Assign the current user as the coordinator
        project.save()
        return super(ProjectCreateView, self).form_valid(form)


class ProjectDetailView(DetailView):
    model = Project
    template_name = "coordinator/project_detail.html"
    context_object_name = "project"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = context["project"]
        context["project_groups"] = [
            {
                "group": group,
                "supervisor": FacultyAssignment.objects.filter(
                    project_group=group, role="SUPERVISOR"
                ).first(),
                "examiners": FacultyAssignment.objects.filter(
                    project_group=group, role="EXAMINER"
                ),
            }
            for group in project.project_groups.all()
        ]
        return context


class ProjectDeleteView(DeleteView):
    model = Project
    template_name = "coordinator/project_confirm_delete.html"
    success_url = reverse_lazy("coordinator-project-list")
    context_object_name = "project"


from .models import Project, ProjectGroup
from .forms import ProjectGroupForm


class GroupListView(LoginRequiredMixin, ListView):
    model = ProjectGroup
    template_name = "coordinator/group_list.html"
    context_object_name = "groups"
    login_url = "/accounts/login/"
    redirect_field_name = "next"

    def get_queryset(self):
        # Filter groups by the project_id passed in the URL
        project_id = self.kwargs.get("project_id")
        return ProjectGroup.objects.filter(project__id=project_id)

    def get_context_data(self, **kwargs):
        # Get the standard context data (which includes the list of groups)
        context = super().get_context_data(**kwargs)
        project_id = self.kwargs.get("project_id")
        project = Project.objects.get(pk=project_id)

        # Calculate total scores for each group
        groups_with_totals = []
        for group in context["groups"]:
            student_totals = calculate_final_grades(group.id)
            groups_with_totals.append((group, student_totals))

        # Update the context with new variables
        context["project"] = project
        context["groups_with_totals"] = groups_with_totals

        return context


class ProjectGroupCreateView(CreateView):
    model = ProjectGroup
    form_class = ProjectGroupForm
    template_name = "coordinator/project_group_form.html"

    def form_valid(self, form):
        form.instance.project = Project.objects.get(pk=self.kwargs["pk"])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            "coordinator-project-detail", kwargs={"pk": self.object.project.pk}
        )


class ProjectGroupUpdateView(UpdateView):
    model = ProjectGroup
    form_class = ProjectGroupForm
    template_name = "coordinator/project_group_form.html"

    def get_success_url(self):
        return reverse_lazy(
            "coordinator-project-detail", kwargs={"pk": self.object.project.pk}
        )


class ProjectGroupDeleteView(DeleteView):
    model = ProjectGroup
    success_url = reverse_lazy("coordinator-project-list")

    def get_queryset(self):
        return super().get_queryset().filter(project__coordinator=self.request.user)

    def delete(self, request, *args, **kwargs):
        """
        Call the delete method on the fetched object and then redirect to the
        success URL.
        """
        self.object = self.get_object()
        # Ensure the user deleting the group is the supervisor of the project
        if self.object.project.coordinator != request.user:
            # Handle what happens if the user is not the supervisor
            from django.http import HttpResponseForbidden

            return HttpResponseForbidden("You are not allowed to delete this group")
        self.object.delete()
        return super().delete(request, *args, **kwargs)


from .forms import FacultyAssignmentForm
from .models import FacultyAssignment
from django.views.generic.edit import FormView


class AssignFacultyView(FormView):
    template_name = "coordinator/assign_faculty.html"
    form_class = FacultyAssignmentForm
    success_url = reverse_lazy(
        "coordinator-project-list"
    )  # Redirect after a successful assignment

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["project_group_id"] = self.kwargs.get(
            "project_group_id"
        )  # Pass project group ID to the template
        return context

    def form_valid(self, form):
        project_group_id = self.kwargs.get("project_group_id")
        project_group = get_object_or_404(ProjectGroup, pk=project_group_id)
        faculty = form.cleaned_data["faculty"]
        role = form.cleaned_data["role"]

        # Check if adding an examiner and if there are already 2 examiners assigned
        if (
            role == "EXAMINER"
            and FacultyAssignment.objects.filter(
                project_group=project_group, role="EXAMINER"
            ).count()
            >= 2
        ):
            messages.error(
                self.request, "Only two examiners can be assigned per group."
            )
            return self.form_invalid(form)

        # Update or create the faculty assignment
        FacultyAssignment.objects.update_or_create(
            project_group=project_group, faculty=faculty, defaults={"role": role}
        )
        messages.success(self.request, "Faculty assigned successfully.")
        return super().form_valid(form)

    def form_invalid(self, form):
        # Optionally handle invalid form cases
        return super().form_invalid(form)


class FacultyProjectListView(LoginRequiredMixin, ListView):
    model = Project
    template_name = "faculty/project_list.html"
    context_object_name = "projects"

    def get_queryset(self):
        # Get all project groups where the current user is assigned as faculty
        faculty_assignments = (
            FacultyAssignment.objects.filter(faculty=self.request.user)
            .select_related("project_group")
            .distinct()
        )
        project_ids = faculty_assignments.values_list(
            "project_group__project_id", flat=True
        )
        return Project.objects.filter(id__in=project_ids).distinct()


class FacultyProjectGroupsView(LoginRequiredMixin, ListView):
    model = ProjectGroup
    template_name = "faculty/project_groups.html"
    context_object_name = "groups"

    def get_queryset(self):
        project_id = self.kwargs.get("project_id")
        # Get groups only for this project and where the current user is a supervisor or examiner
        return ProjectGroup.objects.filter(
            project_id=project_id, assignments__faculty=self.request.user
        ).distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["project"] = Project.objects.get(pk=self.kwargs.get("project_id"))
        return context


class FacultyGroupDetailView(DetailView):
    model = ProjectGroup
    template_name = "faculty/group_detail.html"
    context_object_name = "group"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Optionally add more context data if needed, for instance:
        context["assignments"] = self.object.assignments.all()
        return context


from .models import Logbook, LogbookEntry
from .forms import LogbookEntryForm
from django.shortcuts import redirect, get_object_or_404


class LogbookEntryCreateView(CreateView):
    model = LogbookEntry
    form_class = LogbookEntryForm
    template_name = "supervisor/logbook_entry_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.logbook = get_object_or_404(Logbook, pk=self.kwargs["logbook_id"])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.logbook = self.logbook
        return super().form_valid(form)


class LogbookEntryUpdateView(UpdateView):
    model = LogbookEntry
    form_class = LogbookEntryForm
    template_name = "supervisor/logbook_entry_form.html"


from .models import ProjectGroup, Proposal
from .forms import ProposalForm


class ProposalCreateView(CreateView):
    model = Proposal
    form_class = ProposalForm
    template_name = "supervisor/proposal_form.html"

    def form_valid(self, form):
        # Correctly access self.kwargs to fetch the project_group_id
        project_group_id = self.kwargs.get("project_group_id")  # Correct attribute name
        form.instance.project_group = ProjectGroup.objects.get(pk=project_group_id)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            "supervisor-project-detail",
            kwargs={"pk": self.object.project_group.project.pk},
        )


class ProposalUpdateView(UpdateView):
    model = Proposal
    form_class = ProposalForm
    template_name = "supervisor/proposal_form.html"

    def get_success_url(self):
        return reverse_lazy(
            "project-detail", kwargs={"pk": self.object.project_group.project.pk}
        )


from .models import ProgressReport
from .forms import ProgressReportForm


class ProgressReportCreateView(CreateView):
    model = ProgressReport
    form_class = ProgressReportForm
    template_name = "supervisor/progress_report_form.html"

    def form_valid(self, form):
        project_group_id = self.kwargs.get(
            "project_group_id"
        )  # Ensure this matches your URL conf
        form.instance.project_group = ProjectGroup.objects.get(pk=project_group_id)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            "supervisor-project-detail",
            kwargs={"pk": self.object.project_group.project.pk},
        )


class ProgressReportUpdateView(UpdateView):
    model = ProgressReport
    form_class = ProgressReportForm
    template_name = "pages/form.html"

    def update_success_url(self, form):
        return reverse_lazy("report-page", kwargs={"pk": self.object.project_group.pk})


from .forms import MidProjectReportForm
from .models import MidProjectReport


class MidProjectReportCreateView(CreateView):
    model = MidProjectReport
    form_class = MidProjectReportForm
    template_name = "pages/form.html"

    def form_valid(self, form):
        form.instance.project_group = ProjectGroup.objects.get(pk=self.kwargs["pk"])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("report-page", kwargs={"pk": self.object.project_group.pk})


from .models import MidProjectPresentation
from .forms import MidProjectPresentationForm


class MidProjectPresentationCreateView(CreateView):
    model = MidProjectPresentation
    form_class = MidProjectPresentationForm
    template_name = "pages/form.html"

    def form_valid(self, form):
        form.instance.project_group = ProjectGroup.objects.get(pk=self.kwargs["pk"])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("report-page", kwargs={"pk": self.object.project_group.pk})


from django.views import View
from django.db import transaction
from .forms import InterviewEntryFormSet
from .models import Interview, InterviewEntry
from django.forms import modelformset_factory
from .forms import InterviewEntryForm


def get_interview_entry_formset(extra=0):
    return modelformset_factory(
        InterviewEntry, form=InterviewEntryForm, extra=extra, can_delete=True
    )


def get_logbook_entry_formset(extra=0):
    return modelformset_factory(
        LogbookEntry, form=LogbookEntryForm, extra=extra, can_delete=True
    )


def get_examiner_entry_formset(extra=0):
    return modelformset_factory(
        ExaminerMidProjectPresentationQA,
        form=ExaminerMidProjectPresentationQAForm,
        extra=extra,
        can_delete=True,
    )


from django.db.models import Sum
from decimal import Decimal, getcontext


def calculate_student_totals(project_group_id):
    project_group = ProjectGroup.objects.get(pk=project_group_id)
    students = project_group.students.all()

    student_totals = {}

    for student in students:
        scores = {}

        # Set precision for Decimal operations
        getcontext().prec = 2

        # Interview (10% of total grade)
        interview_score = Decimal(
            InterviewEntry.objects.filter(
                student=student, interview__project_group=project_group
            ).aggregate(total=Sum("score"))["total"]
            or 0
        )
        scores["interview"] = (interview_score / Decimal("100")) * Decimal("10")

        # Logbooks (10% of total grade)
        logbook_scores = LogbookEntry.objects.filter(
            student=student, logbook__project_group=project_group
        ).aggregate(
            total_planning=Sum("planning_score"),
            total_execution=Sum("execution_score"),
            total_functionality=Sum("functionality_score"),
        )
        logbook_total_score = sum(
            Decimal(score) for score in logbook_scores.values() if score is not None
        )
        scores["logbooks"] = (logbook_total_score / Decimal("100")) * Decimal("10")

        # Proposal (10% of total grade)
        proposal = Proposal.objects.filter(project_group=project_group).last()
        if proposal:
            proposal_score = (
                Decimal(proposal.planning_score)
                + Decimal(proposal.execution_score)
                + Decimal(proposal.formatting_score)
                + Decimal(proposal.presentation_score)
            )
            scores["proposal"] = (proposal_score / Decimal("100")) * Decimal("10")

        # Progress Report (20% of total grade)
        progress_report = ProgressReport.objects.filter(
            project_group=project_group
        ).last()
        if progress_report:
            progress_report_score = Decimal(progress_report.formatting_score) + Decimal(
                progress_report.design_process_score
            )
            scores["progress_report"] = (
                progress_report_score / Decimal("100")
            ) * Decimal("20")

        # Mid-Project Report (20% of total grade)
        mid_project_report = MidProjectReport.objects.filter(
            project_group=project_group
        ).last()
        if mid_project_report:
            mid_project_report_score = (
                Decimal(mid_project_report.formatting_score)
                + Decimal(mid_project_report.content_score)
                + Decimal(mid_project_report.design_selection_score)
                + Decimal(mid_project_report.conclusions_score)
            )
            scores["mid_project_report"] = (
                mid_project_report_score / Decimal("100")
            ) * Decimal("20")

        # Mid-Project Presentation (30% of total grade)
        mid_project_presentation = MidProjectPresentation.objects.filter(
            project_group=project_group
        ).last()
        if mid_project_presentation:
            mid_project_presentation_score = (
                Decimal(mid_project_presentation.format_score)
                + Decimal(mid_project_presentation.technical_content_score)
                + Decimal(mid_project_presentation.qa_score)
            )
            scores["mid_project_presentation"] = (
                mid_project_presentation_score / Decimal("100")
            ) * Decimal("30")

        # Calculate total percentage
        total_percentage = sum(scores.values())
        student_totals[student.username] = total_percentage

    return student_totals


def calculate_examiner_totals(project_group_id):
    project_group = ProjectGroup.objects.get(pk=project_group_id)
    students = project_group.students.all()

    student_totals = {}

    # Set precision for Decimal operations
    getcontext().prec = 2

    for student in students:
        scores = {"format": 0, "technical": 0, "report": 0, "qa": 0}

        # Fetch all related presentations for the group
        presentations = ExaminerMidProjectPresentation.objects.filter(
            project_group=project_group
        )

        # Aggregate scores from all examiners
        for presentation in presentations:
            presentation_scores = ExaminerMidProjectPresentationQA.objects.filter(
                presentation=presentation, student=student
            ).aggregate(
                total_format=Sum("presentation__format_score") or 0,
                total_technical=Sum("presentation__technical_content_score") or 0,
                total_report=Sum("presentation__report_score") or 0,
                total_qa=Sum("qa_score") or 0,
            )

            # Calculate score contributions
            scores["format"] += presentation_scores[
                "total_format"
            ]  # Scale proportionally
            scores["technical"] += presentation_scores["total_technical"]
            scores["report"] += presentation_scores["total_report"]
            scores["qa"] += presentation_scores["total_qa"]

        # Sum all scores to get the total percentage contribution from both examiners
        total_examiner_percentage = sum(scores.values())
        student_totals[student.username] = int(total_examiner_percentage)

    return student_totals


from decimal import Decimal, InvalidOperation, getcontext


def calculate_final_grades(project_group_id):
    # Get the total percentages from both functions
    supervisor_scores = calculate_student_totals(project_group_id)
    examiner_scores = calculate_examiner_totals(project_group_id)

    final_grades = {}

    # Combine the scores based on the specified weightings
    for student in supervisor_scores:
        # Get the supervisor score and scale it to 70%
        supervisor_contribution = supervisor_scores[student] * Decimal("0.7")

        # Get the examiner score, normalize it from a 200-point scale to 100, and scale it to 30%
        examiner_contribution = (
            examiner_scores.get(student, 0) / Decimal("2")
        ) * Decimal("0.3")

        # Sum the contributions to get the final grade out of 100
        final_grade = supervisor_contribution + examiner_contribution

        # Store the final grade, rounding to two decimal places
        try:
            final_grade = final_grade.quantize(Decimal("1.00"))
        except InvalidOperation:
            final_grade = final_grade

        final_grades[student] = final_grade

    return final_grades


from django.http import HttpResponseForbidden


class ProjectReportsView(LoginRequiredMixin, View):
    template_name = "faculty/project_reports_form.html"

    def get(self, request, *args, **kwargs):
        project_group_id = self.kwargs.get("project_group_id")
        project_group = get_object_or_404(ProjectGroup, pk=project_group_id)

        # Check if the current user is the supervisor
        if not FacultyAssignment.objects.filter(
            project_group=project_group, faculty=request.user, role="SUPERVISOR"
        ).exists():
            return HttpResponseForbidden("You are not authorized to view this page.")

        latest_proposal = project_group.proposals.order_by("-id").first()
        latest_progress_report = project_group.progress_reports.order_by("-id").first()
        latest_mid_project_report = project_group.mid_project_reports.order_by(
            "-id"
        ).first()
        latest_mid_project_presentation = (
            project_group.mid_project_presentations.order_by("-id").first()
        )

        # Retrieve or create Interview and Logbook instances
        interview, _ = Interview.objects.get_or_create(project_group=project_group)
        logbook, _ = Logbook.objects.get_or_create(project_group=project_group)

        # Handle existing forms and formsets
        students_count = project_group.students.count()
        extra_forms = max(
            0,
            students_count - InterviewEntry.objects.filter(interview=interview).count(),
        )

        InterviewEntryFormSet = get_interview_entry_formset(extra=extra_forms)
        LogbookEntryFormSet = get_logbook_entry_formset(extra=extra_forms)

        logbook_formset = LogbookEntryFormSet(
            queryset=LogbookEntry.objects.filter(logbook=logbook)
        )

        # Calculate total scores for logbook entries
        logbook_totals = {}
        for form in logbook_formset.forms:
            if form.instance.id:
                total = (
                    form.instance.planning_score
                    + form.instance.execution_score
                    + form.instance.functionality_score
                )
                logbook_totals[form.instance.id] = total

        # Calculate subtotals (example: summing hypothetical 'score' fields)
        # Initialize subtotals to 0
        proposal_subtotal = 0
        progress_report_subtotal = 0
        mid_project_report_subtotal = 0
        mid_project_presentation_subtotal = 0

        # Calculate subtotals if the reports exist
        if latest_proposal:
            proposal_subtotal = sum(
                [
                    getattr(latest_proposal, field.name, 0)
                    for field in latest_proposal._meta.fields
                    if "score" in field.name
                ]
            )
        if latest_progress_report:
            progress_report_subtotal = sum(
                [
                    getattr(latest_progress_report, field.name, 0)
                    for field in latest_progress_report._meta.fields
                    if "score" in field.name
                ]
            )
        if latest_mid_project_report:
            mid_project_report_subtotal = sum(
                [
                    getattr(latest_mid_project_report, field.name, 0)
                    for field in latest_mid_project_report._meta.fields
                    if "score" in field.name
                ]
            )
        if latest_mid_project_presentation:
            mid_project_presentation_subtotal = sum(
                [
                    getattr(latest_mid_project_presentation, field.name, 0)
                    for field in latest_mid_project_presentation._meta.fields
                    if "score" in field.name
                ]
            )

        # proposal_subtotal = sum([getattr(latest_proposal, field.name, 0) for field in latest_proposal._meta.fields if 'score' in field.name])
        # progress_report_subtotal = sum([getattr(latest_progress_report, field.name, 0) for field in latest_progress_report._meta.fields if 'score' in field.name])
        # mid_project_report_subtotal = sum([getattr(latest_mid_project_report, field.name, 0) for field in latest_mid_project_report._meta.fields if 'score' in field.name])
        # mid_project_presentation_subtotal = sum([getattr(latest_mid_project_presentation, field.name, 0) for field in latest_mid_project_presentation._meta.fields if 'score' in field.name])

        student_totals = calculate_student_totals(project_group_id)

        context = {
            "project_group": project_group,
            "proposal_form": ProposalForm(instance=latest_proposal, prefix="proposal"),
            "progress_report_form": ProgressReportForm(
                instance=latest_progress_report, prefix="progress"
            ),
            "mid_project_report_form": MidProjectReportForm(
                instance=latest_mid_project_report, prefix="midreport"
            ),
            "mid_project_presentation_form": MidProjectPresentationForm(
                instance=latest_mid_project_presentation, prefix="midpres"
            ),
            "interview_formset": InterviewEntryFormSet(
                queryset=InterviewEntry.objects.filter(interview=interview),
                form_kwargs={"project_group": project_group},
            ),
            "logbook_formset": LogbookEntryFormSet(
                queryset=LogbookEntry.objects.filter(logbook=logbook),
                form_kwargs={"project_group": project_group},
            ),
            "logbook_totals": logbook_totals,
            "proposal_subtotal": proposal_subtotal,
            "progress_report_subtotal": progress_report_subtotal,
            "mid_project_report_subtotal": mid_project_report_subtotal,
            "mid_project_presentation_subtotal": mid_project_presentation_subtotal,
            "student_totals": student_totals,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        project_group_id = self.kwargs.get("project_group_id")
        project_group = get_object_or_404(ProjectGroup, pk=project_group_id)

        # Get or create Interview and Logbook instances
        interview, _ = Interview.objects.get_or_create(project_group=project_group)
        logbook, _ = Logbook.objects.get_or_create(project_group=project_group)

        # Initialize formsets with the POST data
        InterviewEntryFormSet = get_interview_entry_formset()
        LogbookEntryFormSet = get_logbook_entry_formset()

        interview_formset = InterviewEntryFormSet(
            request.POST,
            form_kwargs={"project_group": project_group},
            queryset=InterviewEntry.objects.filter(interview=interview),
        )
        logbook_formset = LogbookEntryFormSet(
            request.POST,
            form_kwargs={"project_group": project_group},
            queryset=LogbookEntry.objects.filter(logbook=logbook),
        )

        # Initialize other forms
        forms = {
            "proposal_form": ProposalForm(
                request.POST, instance=project_group.proposals.last(), prefix="proposal"
            ),
            "progress_report_form": ProgressReportForm(
                request.POST,
                instance=project_group.progress_reports.last(),
                prefix="progress",
            ),
            "mid_project_report_form": MidProjectReportForm(
                request.POST,
                instance=project_group.mid_project_reports.last(),
                prefix="midreport",
            ),
            "mid_project_presentation_form": MidProjectPresentationForm(
                request.POST,
                instance=project_group.mid_project_presentations.last(),
                prefix="midpres",
            ),
        }

        # Inside your POST method, after initializing forms and formsets
        if not all(form.is_valid() for form in forms.values()):
            for form_name, form in forms.items():
                if not form.is_valid():
                    print(
                        f"Errors in {form_name}: {form.errors}"
                    )  # Use print for debugging or logging

        if not interview_formset.is_valid():
            print(f"Errors in interview_formset: {interview_formset.errors}")

        if not logbook_formset.is_valid():
            print(f"Errors in logbook_formset: {logbook_formset.errors}")

        proposal_form = ProposalForm(request.POST, prefix="proposal")
        if proposal_form.is_valid():
            proposal = proposal_form.save(commit=False)
            proposal.project_group = (
                project_group  # Assign the project group explicitly
            )
            proposal.save()
        else:
            print(proposal_form.errors)

        progress_report_form = ProgressReportForm(request.POST, prefix="progress")
        if progress_report_form.is_valid():
            progress_report = progress_report_form.save(commit=False)
            progress_report.project_group = project_group
            progress_report.save()
        else:
            print(progress_report_form.errors)

        mid_project_report_form = MidProjectReportForm(request.POST, prefix="midreport")
        if mid_project_report_form.is_valid():
            mid_project_report = mid_project_report_form.save(commit=False)
            mid_project_report.project_group = project_group
            mid_project_report.save()
        else:
            print(mid_project_report_form.errors)

        mid_project_presentation_form = MidProjectPresentationForm(
            request.POST, prefix="midpres"
        )
        if mid_project_presentation_form.is_valid():
            mid_project_presentation = mid_project_presentation_form.save(commit=False)
            mid_project_presentation.project_group = project_group
            mid_project_presentation.save()

        else:
            print(mid_project_presentation_form.errors)

        if (
            all(form.is_valid() for form in forms.values())
            and interview_formset.is_valid()
            and logbook_formset.is_valid()
        ):
            with transaction.atomic():
                # for form in forms.values():
                #     form.save()

                # Save the interview entries
                interview_entries = interview_formset.save(commit=False)
                for entry in interview_entries:
                    entry.interview = interview
                    entry.save()
                interview_formset.save_m2m()

                # Save the logbook entries
                logbook_entries = logbook_formset.save(commit=False)
                for entry in logbook_entries:
                    entry.logbook = logbook
                    entry.save()
                logbook_formset.save_m2m()

                return redirect(
                    reverse(
                        "faculty-project-groups",
                        kwargs={"project_id": project_group.project_id},
                    )
                )
        # If forms are not valid, re-render the page with the context containing form errors
        context = {
            **forms,
            "project_group": project_group,
            "interview_formset": interview_formset,
            "logbook_formset": logbook_formset,
        }
        return render(request, self.template_name, context)


def get_examiner_presentation_qa_formset(extra_forms, project_group):
    return modelformset_factory(
        ExaminerMidProjectPresentationQA,
        form=ExaminerMidProjectPresentationQAForm,
        extra=extra_forms,
        can_delete=True,
    )


class ExaminarProjectReportsView(LoginRequiredMixin, View):
    template_name = "faculty/examiner_project_reports_form.html"

    def get(self, request, *args, **kwargs):
        project_group_id = self.kwargs.get("project_group_id")
        project_group = get_object_or_404(ProjectGroup, pk=project_group_id)

        if not FacultyAssignment.objects.filter(
            project_group=project_group, faculty=request.user, role="EXAMINER"
        ).exists():
            return HttpResponseForbidden("You are not authorized to view this page.")

        latest_presentation = (
            project_group.examiner_mid_project_presentations.filter(
                examiner=request.user
            )
            .order_by("-id")
            .first()
        )

        students = project_group.students.all()
        existing_qas = (
            ExaminerMidProjectPresentationQA.objects.filter(
                presentation=latest_presentation
            )
            if latest_presentation
            else ExaminerMidProjectPresentationQA.objects.none()
        )
        extra_forms = max(0, students.count() - existing_qas.count())

        # Create the formset class with the correct number of extra forms
        ExaminerPresentationQAFormSet = get_examiner_presentation_qa_formset(
            extra_forms, project_group
        )

        # Instantiate the formset with queryset
        qa_formset = ExaminerPresentationQAFormSet(
            queryset=existing_qas, form_kwargs={"project_group": project_group}
        )

        presentation_form = ExaminerMidProjectPresentationForm(
            instance=latest_presentation
        )

        context = {
            "project_group": project_group,
            "presentation_form": presentation_form,
            "qa_formset": qa_formset,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        project_group_id = self.kwargs.get("project_group_id")
        project_group = get_object_or_404(ProjectGroup, pk=project_group_id)

        latest_presentation = (
            project_group.examiner_mid_project_presentations.filter(
                examiner=request.user
            )
            .order_by("-id")
            .first()
        )

        # Calculate the number of extra forms needed
        students = project_group.students.all()
        existing_qas = (
            ExaminerMidProjectPresentationQA.objects.filter(
                presentation=latest_presentation
            )
            if latest_presentation
            else ExaminerMidProjectPresentationQA.objects.none()
        )
        extra_forms = max(0, students.count() - existing_qas.count())

        # Create the formset class with the correct number of extra forms
        ExaminerPresentationQAFormSet = get_examiner_presentation_qa_formset(
            extra_forms, project_group
        )

        # Now instantiate the formset with the POST data
        qa_formset = ExaminerPresentationQAFormSet(
            request.POST,
            queryset=existing_qas,
            form_kwargs={"project_group": project_group},
        )

        presentation_form = ExaminerMidProjectPresentationForm(
            request.POST, instance=latest_presentation
        )

        if presentation_form.is_valid() and qa_formset.is_valid():
            with transaction.atomic():
                presentation = presentation_form.save(commit=False)
                presentation.examiner = request.user
                presentation.project_group = project_group
                presentation.save()

                # Save the Q&A entries, making sure each student has an entry
                qa_entries = qa_formset.save(commit=False)
                for entry in qa_entries:
                    entry.presentation = presentation
                    entry.save()
                qa_formset.save_m2m()

                return redirect(
                    reverse(
                        "faculty-project-groups",
                        kwargs={"project_id": project_group.project_id},
                    )
                )

        # Re-render the page with errors if any forms are invalid
        context = {
            "presentation_form": presentation_form,
            "qa_formset": qa_formset,
            "project_group": project_group,
        }
        return render(request, self.template_name, context)


class ExaminerProjectGroupListView(LoginRequiredMixin, ListView):
    model = ProjectGroup
    template_name = "examiner/project_group_list.html"
    context_object_name = "groups"
    login_url = "/accounts/login/"
    redirect_field_name = "next"

    def get_queryset(self):
        # Get the project ID from the URL
        project_id = self.kwargs.get("project_id")

        # Get all project groups of the specific project where the current user is assigned as an examiner
        return ProjectGroup.objects.filter(
            project_id=project_id, examiner_assignments__examiners=self.request.user
        ).distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["project"] = Project.objects.get(pk=self.kwargs.get("project_id"))
        return context


from django.contrib.auth.mixins import UserPassesTestMixin


class StudentProjectReportsView(View):
    template_name = "students/project_reports_form.html"

    def get(self, request, *args, **kwargs):
        project_group_id = self.kwargs.get("project_group_id")
        project_group = get_object_or_404(ProjectGroup, pk=project_group_id)

        latest_proposal = project_group.proposals.order_by("-id").first()
        latest_progress_report = project_group.progress_reports.order_by("-id").first()
        latest_mid_project_report = project_group.mid_project_reports.order_by(
            "-id"
        ).first()
        latest_mid_project_presentation = (
            project_group.mid_project_presentations.order_by("-id").first()
        )

        interview, _ = Interview.objects.get_or_create(project_group=project_group)
        logbook, _ = Logbook.objects.get_or_create(project_group=project_group)

        # Create formsets but make them readonly
        interview_formset = InterviewEntryFormSet(
            queryset=InterviewEntry.objects.filter(interview=interview),
            form_kwargs={"project_group": project_group, "readonly": True},
        )
        logbook_formset = LogbookEntryFormSet(
            queryset=LogbookEntry.objects.filter(logbook=logbook),
            form_kwargs={"project_group": project_group, "readonly": True},
        )

        # Calculate total scores for logbook entries
        logbook_totals = {}
        for form in logbook_formset.forms:
            if form.instance.id:
                total = (
                    form.instance.planning_score
                    + form.instance.execution_score
                    + form.instance.functionality_score
                )
                logbook_totals[form.instance.id] = total

        # Initialize subtotals to 0
        proposal_subtotal = 0
        progress_report_subtotal = 0
        mid_project_report_subtotal = 0
        mid_project_presentation_subtotal = 0

        # Calculate subtotals if the reports exist
        if latest_proposal:
            proposal_subtotal = sum(
                [
                    getattr(latest_proposal, field.name, 0)
                    for field in latest_proposal._meta.fields
                    if "score" in field.name
                ]
            )
        if latest_progress_report:
            progress_report_subtotal = sum(
                [
                    getattr(latest_progress_report, field.name, 0)
                    for field in latest_progress_report._meta.fields
                    if "score" in field.name
                ]
            )
        if latest_mid_project_report:
            mid_project_report_subtotal = sum(
                [
                    getattr(latest_mid_project_report, field.name, 0)
                    for field in latest_mid_project_report._meta.fields
                    if "score" in field.name
                ]
            )
        if latest_mid_project_presentation:
            mid_project_presentation_subtotal = sum(
                [
                    getattr(latest_mid_project_presentation, field.name, 0)
                    for field in latest_mid_project_presentation._meta.fields
                    if "score" in field.name
                ]
            )

        student_totals = calculate_student_totals(project_group_id)

        context = {
            "project_group": project_group,
            "proposal_form": ProposalForm(instance=latest_proposal, prefix="proposal"),
            "progress_report_form": ProgressReportForm(
                instance=latest_progress_report, prefix="progress"
            ),
            "mid_project_report_form": MidProjectReportForm(
                instance=latest_mid_project_report, prefix="midreport"
            ),
            "mid_project_presentation_form": MidProjectPresentationForm(
                instance=latest_mid_project_presentation, prefix="midpres"
            ),
            "interview_formset": interview_formset,
            "logbook_formset": logbook_formset,
            "logbook_totals": logbook_totals,
            "proposal_subtotal": proposal_subtotal,
            "progress_report_subtotal": progress_report_subtotal,
            "mid_project_report_subtotal": mid_project_report_subtotal,
            "mid_project_presentation_subtotal": mid_project_presentation_subtotal,
            "student_totals": student_totals,
        }
        return render(request, self.template_name, context)


class StudentProjectListView(LoginRequiredMixin, ListView):
    model = Project
    template_name = "students/student_project_list.html"
    context_object_name = "projects"

    def get_queryset(self):
        return Project.objects.filter(
            project_groups__students=self.request.user
        ).distinct()

    login_url = "/accounts/login/"  # Redirect to login if not authenticated
    redirect_field_name = "next"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project_groups = {
            group.project_id: group.id
            for group in ProjectGroup.objects.filter(students=self.request.user)
        }
        context["project_groups"] = project_groups
        return context


class StudentGroupDetailsView(LoginRequiredMixin, DetailView):
    model = ProjectGroup
    template_name = "students/student_group_details.html"
    context_object_name = "group"

    def get_object(self):
        return get_object_or_404(
            ProjectGroup, id=self.kwargs["group_id"], students=self.request.user
        )

    def get_context_data(self, **kwargs):
        context = super(StudentGroupDetailsView, self).get_context_data(**kwargs)
        # Calculate final grades for the group
        project_group_id = self.get_object().id
        context["final_grades"] = calculate_final_grades(project_group_id)
        return context


import openpyxl
from openpyxl import load_workbook
from django.http import HttpResponse
from io import BytesIO
from django.utils import timezone


def export_supervisor_assessment(request, project_group_id):
    template_path = "home/templates/SupervisorTemplate.xlsx"
    file_path = "home/templates/SupervisorTemplate.xlsx"

    # Load the workbook and select the active worksheet
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    # Save the workbook with the changes
    workbook.save(filename=file_path)
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    current_date = timezone.now().date()
    ws[f"G4"] = current_date

    # Fetch the project group from the database
    project_group = get_object_or_404(ProjectGroup, pk=project_group_id)
    project = project_group.project

    supervisor_assignment = FacultyAssignment.objects.filter(
        project_group=project_group, role="SUPERVISOR"
    ).first()

    ws["B4"] = "Group #: " + project_group.name
    ws["B5"] = "Project Name: " + project.name

    ws["B6"] = "Supervisor name/s: " + supervisor_assignment.faculty.get_username()

    students = project_group.students.all()
    student_cells = ["B15", "B16", "B17", "B18", "B19"]
    for student, cell in zip(students, student_cells):
        ws[cell] = student.get_username()

    logbook = Logbook.objects.filter(project_group=project_group).first()
    if logbook:
        logbook_entries = LogbookEntry.objects.filter(logbook=logbook)
        columns = ["D", "E", "F", "G", "H"]

        for entry, column in zip(logbook_entries, columns):
            ws[f"{column}27"] = entry.planning_score
            ws[f"{column}28"] = entry.execution_score
            ws[f"{column}29"] = entry.functionality_score

    interview = Interview.objects.filter(project_group=project_group).first()
    if interview:
        interview_entries = InterviewEntry.objects.filter(interview=interview)
        columns = ["D", "E", "F", "G", "H"]

        for entry, column in zip(interview_entries, columns):
            ws[f"{column}25"] = entry.score

    latest_proposal = (
        Proposal.objects.filter(project_group=project_group).order_by("-id").first()
    )
    if latest_proposal:
        ws["F33"] = latest_proposal.planning_score
        ws["F34"] = latest_proposal.execution_score
        ws["F35"] = latest_proposal.formatting_score
        ws["F36"] = latest_proposal.presentation_score

        total_proposal = sum(
            [
                latest_proposal.planning_score,
                latest_proposal.execution_score,
                latest_proposal.formatting_score,
                latest_proposal.presentation_score,
            ]
        )
        ws["F37"] = total_proposal

    latest_progress = (
        ProgressReport.objects.filter(project_group=project_group)
        .order_by("-id")
        .first()
    )
    if latest_progress:
        ws["F39"] = latest_progress.formatting_score
        ws["F40"] = latest_progress.design_process_score

        total_progress = sum(
            [latest_progress.formatting_score, latest_progress.design_process_score]
        )
        ws["F41"] = total_progress

    midReport = (
        MidProjectReport.objects.filter(project_group=project_group)
        .order_by("-id")
        .first()
    )
    if midReport:
        ws["F43"] = midReport.formatting_score
        ws["F44"] = midReport.content_score
        ws["F45"] = midReport.design_selection_score
        ws["F46"] = midReport.conclusions_score

        totalMidReport = sum(
            [
                midReport.formatting_score,
                midReport.content_score,
                midReport.design_selection_score,
                midReport.conclusions_score,
            ]
        )
        ws["F47"] = totalMidReport

    midPresentation = (
        MidProjectPresentation.objects.filter(project_group=project_group)
        .order_by("-id")
        .first()
    )
    if midPresentation:
        ws["F49"] = midPresentation.format_score
        ws["F50"] = midPresentation.technical_content_score
        ws["F51"] = midPresentation.qa_score

        totalMidPresentation = sum(
            [
                midPresentation.format_score,
                midPresentation.technical_content_score,
                midPresentation.qa_score,
            ]
        )
        ws["F52"] = totalMidPresentation

    ws["B10"] = midPresentation.comment

    for student, col in zip(students, columns):
        if logbook:
            entries = LogbookEntry.objects.filter(logbook=logbook, student=student)
            logbook_subtotal = sum(
                entry.planning_score + entry.execution_score + entry.functionality_score
                for entry in entries
            )
            ws[f"{col}31"] = logbook_subtotal  # Storing subtotal for logbook entries

        if interview:
            interview_entries = InterviewEntry.objects.filter(
                interview=interview, student=student
            )
            interview_score = sum(entry.score for entry in interview_entries)

        # Total score calculation considering each component's weight if applicable
        total_score = (
            logbook_subtotal
            + total_proposal
            + total_progress
            + totalMidReport
            + totalMidPresentation
            + interview_score
        )
        total_score = total_score / 600
        total_score = total_score * 100

        ws[f"{col}53"] = total_score

    # Save the updated workbook to a BytesIO stream
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    # Prepare the response with the modified Excel file
    response = HttpResponse(
        virtual_workbook.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{project_group.name}_Grades.xlsx"'
    )
    return response


import zipfile


def export_examiner_assesment(request, project_group_id):
    project_group = get_object_or_404(ProjectGroup, pk=project_group_id)
    examiners = FacultyAssignment.objects.filter(
        project_group=project_group, role="EXAMINER"
    )

    # Create a BytesIO object to hold the ZIP file
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for assignment in examiners:
            examiner = assignment.faculty
            template_path = "home/templates/ExaminerTemplate.xlsx"
            file_path = "home/templates/ExaminerTemplate.xlsx"

            # Load the workbook and select the active worksheet
            workbook = load_workbook(filename=file_path)
            worksheet = workbook.active

            # Save the workbook with the changes
            workbook.save(filename=file_path)
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            # Load the workbook and select the active worksheet
            wb = load_workbook(filename=file_path)
            ws = wb.active

            current_date = timezone.now().date()
            ws["G2"] = current_date

            # Setting basic information
            ws["B3"] = project_group.name
            ws["B4"] = assignment.faculty.get_username()
            ws["B5"] = examiner.get_username()

            # Populate data for the students
            students = project_group.students.all()
            student_cells = ["B13", "B14", "B15", "B16", "B17"]

            mid_project = ExaminerMidProjectPresentation.objects.filter(
                examiner=examiner, project_group=project_group
            ).first()
            if mid_project:
                ws["F20"] = mid_project.format_score
                ws["F21"] = mid_project.technical_content_score
                ws["F23"] = mid_project.report_score
                ws["B8"] = mid_project.comment

            columns = ["D", "E", "F", "G", "H"]
            for student, cell, col in zip(students, student_cells, columns):
                ws[cell] = student.get_username()
                # Fetching QA scores
                qa_entry = ExaminerMidProjectPresentationQA.objects.filter(
                    presentation=mid_project, student=student
                ).first()
                if qa_entry:
                    ws[f"{col}22"] = qa_entry.qa_score

                # Calculate total scores (example calculation, adjust as needed)
                total_score = (
                    mid_project.format_score
                    + mid_project.technical_content_score
                    + mid_project.report_score
                    + qa_entry.qa_score
                    if qa_entry
                    else 0
                )
                ws[f"{col}24"] = total_score

            # Save the workbook to a BytesIO stream
            virtual_workbook = BytesIO()
            wb.save(virtual_workbook)
            virtual_workbook.seek(0)

            zip_file.writestr(
                f"{project_group.name}_{examiner.username}_Assessment.xlsx",
                virtual_workbook.read(),
            )

    # Prepare the response with the ZIP file
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = (
        f'attachment; filename="{project_group.name}_Examiner_Assessments.zip"'
    )

    return response
